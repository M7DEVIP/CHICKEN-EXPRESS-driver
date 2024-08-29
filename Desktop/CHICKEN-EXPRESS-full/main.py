from functools import wraps
import os
import re
from flask import Flask, abort, jsonify, render_template_string, request, render_template, redirect, send_file, url_for, session, send_from_directory, make_response, flash
from datetime import datetime, time, timedelta
import pytz
import requests
import mysql.connector
import socket
from fpdf import FPDF
import arabic_reshaper
from bidi.algorithm import get_display
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties
from collections import Counter
import uuid
import threading
from werkzeug.security import check_password_hash, generate_password_hash
from flask_wtf.csrf import CSRFProtect
from flask_wtf import FlaskForm
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from wtforms import StringField, PasswordField, SubmitField, ValidationError
from wtforms.validators import DataRequired, Length, Optional
from decouple import config, AutoConfig
from dotenv import load_dotenv
from wtforms import StringField, FloatField
from flask_caching import Cache
from waitress import serve
import sys
import io
import logging
from threading import Lock
from queue import Queue
from threading import Thread
from bidi.algorithm import get_display
import bcrypt
from openpyxl.utils import get_column_letter
from authlib.integrations.flask_client import OAuth
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity

app = Flask(__name__)
config = AutoConfig(search_path=os.path.dirname(__file__))
app.config['JWT_SECRET_KEY'] = 'DGHtfyguhhs35346576$#%$65r6fhdd54676o87gvWW$#656y'  # قم بتغيير هذا إلى مفتاح سري قوي
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)
jwt = JWTManager(app)
app.secret_key = os.environ.get('SECRET_KEY', 'default_secret_key')
csrf = CSRFProtect(app)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_COOKIE_SECURE'] = True
cache = Cache(app, config={'CACHE_TYPE': 'simple'})
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
logging.basicConfig(level=logging.DEBUG, format='%(message)s', handlers=[logging.StreamHandler(sys.stdout)])

# تفعيل القيود على محاولات تسجيل الدخول
limiter = Limiter(key_func=get_remote_address)
limiter.init_app(app)



def get_db_connection():
    return mysql.connector.connect(
        host=config('DB_HOST'),
        user=config('DB_USER'),
        port=config('DB_PORT'),
        password=config('DB_PASSWORD'),
        database=config('DB_NAME'),
        ssl_disabled=False  
    )
    
# AUT

def levels_required(*levels):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'level' not in session or session['level'] not in levels:
                flash('ليس لديك الصلاحيات للوصول إلى هذه الصفحة', 'danger')
                return render_template('unauthorized.html')  # عرض صفحة تحتوي على رسالة التنبيه
            return f(*args, **kwargs)
        return decorated_function
    return decorator
    
  # Login pro
  
# إعداد OpenID Connect باستخدام Auth0
oauth = OAuth(app)
oauth.register(
    name='auth0',
    client_id='Iq13afP3bmC2XUbs7IQ6PA4BbR5edcoX',
    client_secret='C1ouILLNXzz7bbp10kFdPgUzyCoYHj4gb7t8MQpkeGvsA8V7U9Rj1RqyZfimOs-g',
    server_metadata_url='https://dev-4mkuqta3k8epy5zv.us.auth0.com',
    client_kwargs={
        'scope': 'openid profile email',
    },
)

class LoginForm(FlaskForm):
    username = StringField('اسم المستخدم', validators=[DataRequired()])
    password = PasswordField('كلمة المرور', validators=[DataRequired()])
    submit = SubmitField('تسجيل الدخول')

@app.route('/')
def index():
    if 'username' in session:
        return redirect(url_for('home'))
    else:
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
@csrf.exempt
def login():
    form = LoginForm()
    error = None
    if request.method == 'POST' and form.validate_on_submit():
        username = form.username.data
        password = form.password.data

        user = get_user_from_db(username)
        if user:
            stored_password = user[1]
            print(f"Stored password from DB: {stored_password}")
            print(f"Entered password: {password}")
            if bcrypt.checkpw(password.encode('utf-8'), stored_password.encode('utf-8')):
                session['username'] = user[2]
                session['level'] = user[3]
                log_action('N/A', datetime.now(), user[2], 'login')
                return redirect(url_for('home'))
            else:
                error = 'اسم المستخدم أو كلمة المرور غير صحيحة!'
        else:
            error = 'اسم المستخدم أو كلمة المرور غير صحيحة!'
        
        flash(error, 'danger')

    return render_template('login.html', form=form, error=error)


def get_user_from_dbapi(username):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT username, password, name, level FROM ceusersm WHERE username = %s", (username,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        return user
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return None


@app.route('/api/login', methods=['POST'])
@csrf.exempt
def api_login():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON data'}), 400

        username = data.get('username')
        password = data.get('password')

        if not username or not password:
            return jsonify({'error': 'Missing username or password'}), 400

        user = get_user_from_dbapi(username)
        
        if user:
            stored_password = user['password']  # افتراض أن كلمة المرور مخزنة تحت المفتاح 'password'
            if bcrypt.checkpw(password.encode('utf-8'), stored_password.encode('utf-8')):
                access_token = create_access_token(identity=username)
                return jsonify({'status': 'success', 'name': user['name'], 'access_token': access_token}), 200  # افتراض أن اسم المستخدم مخزن تحت المفتاح 'name'
            else:
                return jsonify({'error': 'Invalid username or password!'}), 401
        else:
            return jsonify({'error': 'Invalid username or password!'}), 401
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/orders', methods=['GET'])
@jwt_required()
def get_orders_data():
    try:
        current_user = get_jwt_identity()
        db = get_db_connection()
        cursor = db.cursor()
        today = datetime.today().strftime('%Y-%m-%d')
        query = """
        SELECT ordercode, Cphone, `order`, location, locdesc, oroutinzone, ordtocost 
        FROM ceorder 
        WHERE ordtracking = 'تم التجهيز' AND DATE(`date`) = %s
        """
        cursor.execute(query, (today,))
        rows = cursor.fetchall()

        data1 = []
        for row in rows:
            order_code = row[0]
            order_phone = row[1]
            order_order = row[2]
            order_address = row[3]
            order_address_dec = row[4]
            order_zone = row[5]
            order_total_cost = row[6]

            # تحليل تفاصيل الوجبات
            parsed_order = parse_meal_details(order_order)

            data1.append({
                'ordercode': order_code,
                'phone': order_phone,
                'order': parsed_order,
                'address': order_address,
                'address_dec': order_address_dec,
                'zone': order_zone,
                'total_cost': order_total_cost
            })

        cursor.close()
        db.close()

        return jsonify(data1)
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return jsonify([])




@app.route('/callback')
def callback():
    token = oauth.auth0.authorize_access_token()
    user_info = oauth.auth0.parse_id_token(token)
    session['user'] = user_info
    session['username'] = user_info['name']  # إضافة اسم المستخدم إلى الجلسة
    session['level'] = 'admin' if 'admin' in user_info['roles'] else 'user'  # افتراضًا أن الأدوار جزء من المعلومات المسترجعة
    return redirect(url_for('home'))

@app.route('/logout')
def logout():
    session.pop('user', None)
    session.pop('username', None)
    session.pop('level', None)
    return redirect(oauth.auth0.api_base_url + '/v2/logout?returnTo=' + url_for('login', _external=True) + '&client_id=' + oauth.auth0.client_id)

@app.route('/home')
def home():
    if 'username' in session:
        username = session['username']
        customer_data = get_customers_data()
        order_data = get_orders_data()
        auth = 1 if session['level'] == 'admin' else 0
        return render_template('home.html', username=username, customer_data=customer_data, order_data=order_data, auth=auth)
    else:
        return redirect(url_for('login'))


def get_user_from_db(username):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT username, password, name, level FROM ceusersm WHERE username = %s", (username,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        return user
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return None


@app.route('/add_user', methods=['POST'])
@levels_required('admin', 'editor')
@csrf.exempt  # في حال كنت تريد استثناء بعض المسارات من CSRF
def add_user():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    name = data.get('name')
    level = data.get('level')
    added_by = session.get('username')

    if not added_by:
        return jsonify({'message': 'لم يتم تسجيل المستخدم الذي قام بالإضافة'}), 403

    # تحقق من صحة البيانات المدخلة
    if not username or not password or not name or not level:
        return jsonify({'message': 'يجب ملء جميع الحقول'}), 400

    # تشفير كلمة المرور
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO ceusersm (username, password, name, level, added_by) VALUES (%s, %s, %s, %s, %s)",
            (username, hashed_password, name, level, added_by)
        )
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'message': 'تم إضافة المستخدم بنجاح'})
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return jsonify({'message': 'حدث خطأ أثناء إضافة المستخدم'}), 500


def log_action(order_code, action_time, username, action_type):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        sql = "INSERT INTO aclog (order_code, action_time, username, action_type) VALUES (%s, %s, %s, %s)"
        values = (order_code, action_time, username, action_type)
        cursor.execute(sql, values)
        db.commit()
        cursor.close()
        db.close()
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        
# عرض لووووووووووج

@app.route('/get-logs', methods=['GET'])
def get_logs():
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))
        offset = (page - 1) * per_page

        order_code = request.args.get('order_code', '')
        username = request.args.get('username', '')
        action_type = request.args.get('action_type', '')
        action_time_from = request.args.get('action_time_from', '')
        action_time_to = request.args.get('action_time_to', '')

        query = "SELECT order_code, action_time, username, action_type FROM aclog WHERE 1=1"
        params = []

        if order_code:
            query += " AND order_code LIKE %s"
            params.append(f"%{order_code}%")
        if username:
            query += " AND username LIKE %s"
            params.append(f"%{username}%")
        if action_type:
            query += " AND action_type LIKE %s"
            params.append(f"%{action_type}%")
        if action_time_from:
            query += " AND action_time >= %s"
            params.append(action_time_from)
        if action_time_to:
            query += " AND action_time <= %s"
            params.append(action_time_to)

        query += " ORDER BY action_time DESC LIMIT %s OFFSET %s"
        params.extend([per_page, offset])

        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        cursor.execute(query, params)
        logs = cursor.fetchall()
        cursor.execute("SELECT COUNT(*) as total FROM aclog WHERE 1=1")
        total_logs = cursor.fetchone()['total']
        cursor.close()
        db.close()

        return jsonify({
            'draw': request.args.get('draw', 1),
            'recordsTotal': total_logs,
            'recordsFiltered': total_logs,
            'data': logs
        })
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return jsonify({
            'draw': request.args.get('draw', 1),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': []
        })


def get_orders_data_time(period='daily', start_date=None, end_date=None):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = """
        SELECT `date`, `time`
        FROM ceorder
        """
        
        date_filter = {
            'daily': 'CURDATE()',
            'weekly': 'DATE_SUB(CURDATE(), INTERVAL 1 WEEK)',
            'monthly': 'DATE_SUB(CURDATE(), INTERVAL 1 MONTH)'
        }.get(period, None)
        
        if period == 'custom' and start_date and end_date:
            date_filter = f"'{start_date}' AND '{end_date}'"
        
        if date_filter:
            query += f" WHERE `date` >= {date_filter}"

        cursor.execute(query)
        rows = cursor.fetchall()

        orders_per_hour = Counter()

        for row in rows:
            order_date = row[0]
            order_time = row[1]
            order_datetime_str = f"{order_date} {order_time}"
            order_datetime = datetime.strptime(order_datetime_str, '%Y-%m-%d %H:%M:%S')
            order_hour = order_datetime.hour
            orders_per_hour[order_hour] += 1

        cursor.close()
        db.close()

        orders_data = [orders_per_hour[hour] for hour in range(24)]
        return orders_data
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return [0] * 24

@app.route('/get_orders_data_time', methods=['GET'])
def get_orders_data_time_route():
    period = request.args.get('period', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_orders_data_time(period, start_date, end_date)
    return jsonify(data)


def get_top_meals(period='daily', start_date=None, end_date=None):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = """
        SELECT `order`
        FROM ceorder
        """
        
        date_filter = {
            'daily': 'CURDATE()',
            'weekly': 'DATE_SUB(CURDATE(), INTERVAL 1 WEEK)',
            'monthly': 'DATE_SUB(CURDATE(), INTERVAL 1 MONTH)'
        }.get(period, None)
        
        if period == 'custom' and start_date and end_date:
            date_filter = f"'{start_date}' AND '{end_date}'"
        
        if date_filter:
            query += f" WHERE `date` >= {date_filter}"

        cursor.execute(query)
        rows = cursor.fetchall()

        meal_counter = Counter()

        for row in rows:
            order_details = row[0]
            meals = order_details.split('|')
            for meal in meals:
                meal_name = meal.split(' x ')[1].split('(')[0].strip()
                if meal_name:
                    meal_counter[meal_name] += 1

        top_meals = meal_counter.most_common(6)
        cursor.close()
        db.close()
        return top_meals
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return []

@app.route('/get_top_meals', methods=['GET'])
@levels_required('admin', 'editor')
def get_top_meals_route():
    period = request.args.get('period', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_top_meals(period, start_date, end_date)
    return jsonify(data)


def get_orders_status_today(period='daily', start_date=None, end_date=None):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = """
        SELECT `statu`, COUNT(*)
        FROM ceorder
        """
        
        date_filter = {
            'daily': 'CURDATE()',
            'weekly': 'DATE_SUB(CURDATE(), INTERVAL 1 WEEK)',
            'monthly': 'DATE_SUB(CURDATE(), INTERVAL 1 MONTH)'
        }.get(period, None)
        
        if period == 'custom' and start_date and end_date:
            date_filter = f"'{start_date}' AND '{end_date}'"
        
        if date_filter:
            query += f" WHERE `date` >= {date_filter}"

        query += " GROUP BY `statu`"
        cursor.execute(query)
        rows = cursor.fetchall()

        status_counts = {status: count for status, count in rows}
        cursor.close()
        db.close()
        return status_counts
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return {}

@app.route('/get_orders_status_today', methods=['GET'])
@levels_required('admin', 'editor')
def get_orders_status_today_route():
    period = request.args.get('period', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_orders_status_today(period, start_date, end_date)
    return jsonify(data)

def get_daily_revenue(period='daily', start_date=None, end_date=None):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = """
        SELECT SUM(ordtocost)
        FROM ceorder
        WHERE `statu` = 'Done'
        """
        
        date_filter = {
            'daily': 'CURDATE()',
            'weekly': 'DATE_SUB(CURDATE(), INTERVAL 1 WEEK)',
            'monthly': 'DATE_SUB(CURDATE(), INTERVAL 1 MONTH)'
        }.get(period, None)
        
        if period == 'custom' and start_date and end_date:
            date_filter = f"'{start_date}' AND '{end_date}'"
        
        if date_filter:
            query += f" AND `date` >= {date_filter}"

        cursor.execute(query)
        total_revenue = cursor.fetchone()[0] or 0
        cursor.close()
        db.close()
        return total_revenue
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return 0

@app.route('/get_daily_revenue', methods=['GET'])
@levels_required('admin', 'editor')
def get_daily_revenue_route():
    period = request.args.get('period', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_daily_revenue(period, start_date, end_date)
    return jsonify(data)

def get_top_customers(period='monthly', start_date=None, end_date=None):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = """
        SELECT Cname, Cphone, COUNT(*) as orders
        FROM ceorder
        WHERE statu = 'Done'
        """
        
        date_filter = {
            'daily': 'CURDATE()',
            'weekly': 'DATE_SUB(CURDATE(), INTERVAL 1 WEEK)',
            'monthly': 'DATE_SUB(CURDATE(), INTERVAL 1 MONTH)'
        }.get(period, None)
        
        if period == 'custom' and start_date and end_date:
            date_filter = f"'{start_date}' AND '{end_date}'"
        
        if date_filter:
            query += f" AND `date` >= {date_filter}"

        query += """
        GROUP BY Cname, Cphone
        ORDER BY orders DESC
        LIMIT 10
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        top_customers = [{'name': row[0], 'phone': row[1], 'orders': row[2]} for row in rows]
        return top_customers
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return []

@app.route('/get_top_customers', methods=['GET'])
@levels_required('admin', 'editor')
def get_top_customers_route():
    period = request.args.get('period', 'monthly')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_top_customers(period, start_date, end_date)
    return jsonify(data)

@app.route('/get_order_tracking', methods=['GET'])
def get_order_tracking():
    period = request.args.get('period', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    date_filter = {
        'daily': 'CURDATE()',
        'weekly': 'DATE_SUB(CURDATE(), INTERVAL 1 WEEK)',
        'monthly': 'DATE_SUB(CURDATE(), INTERVAL 1 MONTH)'
    }.get(period, None)
    
    if period == 'custom' and start_date and end_date:
        date_filter = f"'{start_date}' AND '{end_date}'"

    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = f"""
            SELECT ordercode, 
                   orddrivename,
                   CASE WHEN statu = 'Lost' THEN wlost ELSE ordtracking END AS ordtracking,
                   TIMESTAMPDIFF(SECOND, order_preparing_start_time, order_preparing_end_time) AS preparing_duration,
                   TIMESTAMPDIFF(SECOND, order_preparing_end_time, order_delivery_start_time) AS driver_wait_duration,
                   TIMESTAMPDIFF(SECOND, order_delivery_start_time, order_delivered_time) AS delivery_duration,
                   TIMESTAMPDIFF(SECOND, order_preparing_start_time, order_delivered_time) AS total_order_duration,
                   statu,
                   TIME_FORMAT(time, '%H:%i:%s') AS time
            FROM ceorder
            WHERE `date` >= {date_filter}
            ORDER BY
                CASE WHEN statu = 'New' THEN 1
                     WHEN statu = 'Lost' THEN 2
                     WHEN statu = 'Done' THEN 3
                END, order_preparing_start_time
        """
        cursor.execute(query)
        data = cursor.fetchall()
        cursor.close()
        db.close()

        columns = ['ordercode', 'orddrivename', 'ordtracking', 'preparing_duration', 'driver_wait_duration', 'delivery_duration', 'total_order_duration', 'statu', 'time']
        df = pd.DataFrame(data, columns=columns)

        # تحويل الفترات الزمنية إلى صيغة مفهومة
        def format_duration(seconds):
            if pd.isna(seconds) or seconds is None:
                return ''
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            seconds = int(seconds % 60)
            return f"{hours:02}:{minutes:02}:{seconds:02}"

        df['preparing_duration'] = df['preparing_duration'].apply(format_duration)
        df['driver_wait_duration'] = df['driver_wait_duration'].apply(format_duration)
        df['delivery_duration'] = df['delivery_duration'].apply(format_duration)
        df['total_order_duration'] = df['total_order_duration'].apply(format_duration)

        result = df.to_dict(orient='records')
        return jsonify(result)
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return jsonify({'error': 'حدث خطأ أثناء جلب البيانات'}), 500


@app.route('/get_delay_statistics', methods=['GET'])
def get_delay_statistics():
    period = request.args.get('period', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    date_filter = {
        'daily': 'CURDATE()',
        'weekly': 'DATE_SUB(CURDATE(), INTERVAL 1 WEEK)',
        'monthly': 'DATE_SUB(CURDATE(), INTERVAL 1 MONTH)'
    }.get(period, None)
    
    if period == 'custom' and start_date and end_date:
        date_filter = f"'{start_date}' AND '{end_date}'"

    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = f"""
            SELECT
                SUM(TIMESTAMPDIFF(SECOND, order_preparing_start_time, order_preparing_end_time)) AS total_preparing_duration,
                SUM(TIMESTAMPDIFF(SECOND, order_preparing_end_time, order_delivery_start_time)) AS total_driver_wait_duration,
                SUM(TIMESTAMPDIFF(SECOND, order_delivery_start_time, order_delivered_time)) AS total_delivery_duration
            FROM ceorder
            WHERE `date` >= {date_filter}
        """
        cursor.execute(query)
        data = cursor.fetchone()
        cursor.close()
        db.close()

        result = {
            'preparing_duration': data[0] if data[0] is not None else 0,
            'driver_wait_duration': data[1] if data[1] is not None else 0,
            'delivery_duration': data[2] if data[2] is not None else 0
        }
        return jsonify(result)
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return jsonify({'error': 'حدث خطأ أثناء جلب البيانات'}), 500

# دوال الطباعة
def get_order_details(ordercode):
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT Cname, Cphone, location, note, `date`, `time`, ordtocost, oroutinzone, `order` FROM ceorder WHERE ordercode = %s", (ordercode,))
        order = cursor.fetchone()
        cursor.close()
        db.close()
        return order
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return None
        


def get_customers_data():
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = "SELECT Cname, Cphone, Cgender, Caddress, Corder, Cnote FROM cecustomer"
        cursor.execute(query)
        rows = cursor.fetchall()

        data = []
        for row in rows:
            customer_name = row[0]
            customer_phone = row[1]
            customer_gender = row[2]
            customer_address = row[3]
            customer_order = row[4]
            customer_note = row[5]
            data.append({
                'Cname': customer_name,
                'Cphone': customer_phone,
                'Cgender': customer_gender,
                'Caddress': customer_address,
                'Corder': customer_order,
                'Cnote': customer_note
            })

        cursor.close()
        db.close()

        return data
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return []

@app.route('/get_customer_data', methods=['GET'])
@levels_required('admin', 'editor')
def get_customer_data():
    phone = request.args.get('phone')
    data = get_customers_data()
    customer = next((item for item in data if item['Cphone'] == phone), None)
    if customer:
        return jsonify(customer)
    else:
        return jsonify({'error': 'Customer not found'}), 404

def get_orders_data():
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = "SELECT ordercode, Cphone, Cname, `order`, location, `date`, `time`, statu, note, oroutinzone, orddiscnt, ordemname, ordtocost FROM ceorder"
        cursor.execute(query)
        rows = cursor.fetchall()

        data1 = []
        for row in rows:
            order_code = row[0]
            order_phone = row[1]
            order_name = row[2]
            order_order = row[3]
            order_address = row[4]
            order_date = row[5]
            order_time = row[6]
            order_status = row[7]
            order_note = row[8]
            order_zone = row[9]
            order_discount = row[10]
            order_employee_name = row[11]
            order_total_cost = row[12]
            data1.append({
                'ordercode': order_code,
                'phone': order_phone,
                'name': order_name,
                'order': order_order,
                'address': order_address,
                'date': order_date,
                'time': order_time,
                'status': order_status,
                'note': order_note,
                'zone': order_zone,
                'discount': order_discount,
                'employee_name': order_employee_name,
                'total_cost': order_total_cost
            })

        cursor.close()
        db.close()

        return data1
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return []

# تعديل يوزر 
@app.route('/edit_customer', methods=['POST'])
@levels_required('admin', 'editor')
def edit_customer():
    phone = request.form.get('phone')
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = "SELECT Cname, Cphone, Cgender, Caddress, Corder, Cnote FROM cecustomer WHERE Cphone = %s"
        cursor.execute(query, (phone,))
        customer_data = cursor.fetchone()
        if customer_data:
            customer_name = customer_data[0]  
            customer_phone = customer_data[1]
            customer_gender = customer_data[2] 
            customer_address = customer_data[3]  
            customer_orders = customer_data[4]  
            customer_note = customer_data[5]  
            return jsonify({
                "name": customer_name,
                "phone": customer_phone,
                "gender": customer_gender,
                "address": customer_address,
                "orders": customer_orders,
                "note": customer_note
            })
        cursor.close()
        db.close()
        abort(404)
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        abort(500)

@app.route('/update_customer', methods=['POST'])
@levels_required('admin', 'editor')
def update_customer():
    customer_name = request.form.get('name')
    customer_phone = request.form.get('phone')
    customer_gender = request.form.get('gender')
    customer_address = request.form.get('address')
    customer_orders = request.form.get('orders')
    customer_note = request.form.get('note')
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = "UPDATE cecustomer SET Cname = %s, Cgender = %s, Caddress = %s, Corder = %s, Cnote = %s WHERE Cphone = %s"
        cursor.execute(query, (customer_name, customer_gender, customer_address, customer_orders, customer_note, customer_phone))
        db.commit()
        cursor.close()
        db.close()
        return jsonify({"message": "تم تحديث بيانات الزبون بنجاح."})
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return jsonify({"error": "حدث خطأ أثناء تحديث بيانات الزبون."}), 500

@app.route('/add_customer', methods=['POST'])
@levels_required('admin', 'editor')
def add_customer():
    name = request.form.get('name1')
    phone = request.form.get('phone1')
    gender = request.form.get('gender1')
    address = request.form.get('address1')
    orders = request.form.get('orders1')
    note = request.form.get('note1')
    print(name, phone, gender, address, orders, note)
    try:
        db = get_db_connection()
        cursor = db.cursor()
        insert_query = "INSERT INTO cecustomer (Cname, Cphone, Cgender, Caddress, Corder, Cnote) VALUES (%s, %s, %s, %s, %s, %s)"
        values = (name, phone, gender, address, orders, note)
        cursor.execute(insert_query, values)
        db.commit()
        cursor.close()
        db.close()
        return jsonify("تمت إضافة الزبون بنجاح")
    except mysql.connector.Error as error:
        print(f"Database error: {error}")
        return jsonify("حدث خطأ أثناء إضافة الزبون: " + str(error)), 500
    

def get_orders_data2(shift='both'):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        now = datetime.now()
        today_date = now.strftime('%Y-%m-%d')
        previous_day_10pm = (now - timedelta(days=1)).replace(hour=22, minute=0, second=0, microsecond=0).strftime('%Y-%m-%d %H:%M:%S')

        base_query = """
        SELECT ordercode, Cphone, Cname, `order`, location, locdesc, `date`, `time`, statu, note, oroutinzone, 
               orddiscnt, ordemname, ordtocost, ordtracking, orddrivename, wlost 
        FROM ceorder
        WHERE (`date` = %s) OR (`date` = DATE(%s) AND `time` >= %s)
        """

        cursor.execute(base_query, (today_date, previous_day_10pm, previous_day_10pm))
        rows = cursor.fetchall()

        data2 = []
        for row in rows:
            if isinstance(row[7], str):
                order_time = datetime.strptime(row[7], '%H:%M:%S').time()
            elif isinstance(row[7], timedelta):
                order_time = (datetime.min + row[7]).time()
            elif isinstance(row[7], datetime):
                order_time = row[7].time()
            else:
                raise ValueError(f"Unexpected time format: {type(row[7])}")

            order_date = row[6]  # Assuming this is already a datetime.date object

            shift_type = 'evening' if order_time >= time(16, 30) or (order_date == (now - timedelta(days=1)).date() and order_time >= time(22, 0)) else 'morning'
            if shift == 'both' or shift == shift_type:
                data2.append({
                    'ordercode': row[0],
                    'phone': row[1],
                    'name': row[2],
                    'order': row[3],
                    'address': row[4],
                    'address_description': row[5],
                    'date': row[6],
                    'time': row[7],
                    'status': row[8],
                    'note': row[9],
                    'zone': row[10],
                    'discount': row[11],
                    'employee_name': row[12],
                    'total_cost': row[13],
                    'tracking': row[14],
                    'driver_name': row[15],
                    'wlost': row[16]
                })

        cursor.close()
        db.close()

        return data2
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return []
    except ValueError as verr:
        print(f"Value error: {verr}")
        return []


@app.route('/orderwin', methods=['GET'])
def orderwin():
    if 'username' in session:
        username = session['username']
        user_level = session.get('level', 'viewer')  # افتراض المستوى 'viewer' إذا لم يكن محددًا
        shift = request.args.get('shift', 'both')
        order_data = get_orders_data2(shift)
        menu = showMenu()
        return render_template('order.html', username=username, user_level=user_level, order_data=order_data, menu=menu)
    return redirect(url_for('login'))






@app.route('/get_orders_count_by_tracking', methods=['GET'])
def get_orders_count_by_tracking():
    try:
        shift = request.args.get('shift', 'both')
        db = get_db_connection()
        cursor = db.cursor()
        now = datetime.now()
        today_date = now.strftime('%Y-%m-%d')
        previous_day_10pm = (now - timedelta(days=1)).replace(hour=22, minute=0, second=0, microsecond=0)

        base_query = """
        SELECT ordtracking, `time`
        FROM ceorder
        WHERE (`date` = %s OR (`date` = DATE(%s) AND `time` >= %s)) AND statu != 'Lost'
        """

        cursor.execute(base_query, (today_date, previous_day_10pm.strftime('%Y-%m-%d'), previous_day_10pm.strftime('%H:%M:%S')))
        rows = cursor.fetchall()

        tracking_counts = {}

        for tracking, order_time in rows:
            if isinstance(order_time, str):
                order_time = datetime.strptime(order_time, '%H:%M:%S').time()
            elif isinstance(order_time, timedelta):
                order_time = (datetime.min + order_time).time()
            elif isinstance(order_time, datetime):
                order_time = order_time.time()
            else:
                raise ValueError(f"Unexpected time format: {type(order_time)}")

            order_date = (previous_day_10pm.date() if order_time >= time(22, 0) else today_date)
            shift_type = 'evening' if order_time >= time(16, 30) or (order_date == previous_day_10pm.date() and order_time >= time(22, 0)) else 'morning'

            if shift == 'both' or shift == shift_type:
                if tracking is None:
                    tracking = "غير محدد"
                if tracking in tracking_counts:
                    tracking_counts[tracking] += 1
                else:
                    tracking_counts[tracking] = 1

        cursor.close()
        db.close()

        return jsonify(tracking_counts)
    except mysql.connector.Error as err:
        return jsonify({})
    except ValueError as verr:
        print(f"Value error: {verr}")
        return jsonify({})






@app.route('/customerwin')
@levels_required('admin', 'editor')
def customerwin():
    if 'username' in session:
        username = session['username']
        customer_data = get_customers_data()
    return render_template('customer.html', username=username, customer_data=customer_data)

@app.route('/homewin')
def homewin():
    if 'username' in session:
        username = session['username']
        customer_data = get_customers_data()
    return render_template('home.html', username=username, customer_data=customer_data)

@app.route('/adminwin')
@levels_required('admin')
def adminwin():
    if 'username' in session and 'level' in session:
        username = session['username']
        if session['level'] == 'admin':
            return render_template('admin.html', username=session['username'])
        else:
            return "Access Denied", 403
    return render_template('admin.html', username=username)


class OrderForm(FlaskForm):
    order_id = StringField('Order ID', validators=[DataRequired(), Length(max=50)])
    order = StringField('Order', validators=[DataRequired()])
    phone = StringField('Phone', validators=[DataRequired(), Length(max=11)])
    name = StringField('Name', validators=[Optional(), Length(max=100)])
    address = StringField('Address', validators=[DataRequired(), Length(max=255)])
    address_description = StringField('Address Description', validators=[Optional(), Length(max=500)])
    note = StringField('Note', validators=[Optional(), Length(max=500)])
    zone = StringField('Zone', validators=[DataRequired(), Length(max=50)])
    tocost = FloatField('Total Cost', validators=[DataRequired()])
    discount = StringField('Discount', validators=[DataRequired(), Length(max=10)])
    
def validate_discount(form, field):
        value = field.data
        if value.endswith('%'):
            value = value.rstrip('%')
        try:
            float(value)  # محاولة تحويل القيمة إلى عدد للتحقق من صحتها
        except ValueError:
            raise ValidationError("Invalid discount value. It must be a number or a percentage.")

def update_printer_settings_cache():
    global printer_settings_cache
    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute("SELECT ip, port FROM printer_settings LIMIT 1")
        result = cursor.fetchone()
        if result:
            printer_settings_cache = (result[0], result[1])
        else:
            printer_settings_cache = None
            raise Exception("No printer settings found")
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        raise
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

@app.route('/get-printer-settings', methods=['GET'])
@levels_required('admin')
def get_printer_settings_route():
    try:
        settings = get_printer_settings()
        if settings:
            return jsonify({'ip': settings[0], 'port': settings[1]}), 200
        else:
            return jsonify({'error': 'No printer settings found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
def check_printer(ip, port, result):
    try:
        s = socket.create_connection((ip, port), timeout=5)
        s.close()
        result['connected'] = True
    except (socket.timeout, ConnectionRefusedError, socket.error):
        result['connected'] = False
    except Exception as e:
        result['error'] = str(e)

@app.route('/check-printer-status', methods=['GET'])
@levels_required('admin')
def check_printer_status():
    ip = request.args.get('ip')
    port = int(request.args.get('port'))

    result = {}
    thread = threading.Thread(target=check_printer, args=(ip, port, result))
    thread.start()
    thread.join()  # انتظر حتى ينتهي الخيط

    if 'error' in result:
        return jsonify({'error': result['error']}), 500
    else:
        return jsonify({'connected': result['connected']}), 200


@app.route('/save-printer-settings', methods=['POST'])
@levels_required('admin')
def save_printer_settings():
    ip = request.form['ip']
    port = request.form['port']

    try:
        db = get_db_connection()
        cursor = db.cursor()

        # حذف الإعدادات الحالية إذا كانت موجودة
        cursor.execute("DELETE FROM printer_settings")

        # إدخال الإعدادات الجديدة
        cursor.execute("INSERT INTO printer_settings (ip, port) VALUES (%s, %s)", (ip, port))
        db.commit()

        # تحديث الكاش بعد حفظ الإعدادات الجديدة
        update_printer_settings_cache()

        return 'تم حفظ إعدادات الطابعة بنجاح', 200
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return 'حدث خطأ أثناء حفظ إعدادات الطابعة', 500
    except Exception as e:
        print(f"Unexpected error: {e}")
        return 'حدث خطأ غير متوقع أثناء حفظ إعدادات الطابعة', 500
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


pdf_lock = Lock()

def create_pdfs_from_text(text, pdf_path, font_path, lines_per_page=15):
    reshaped_text = arabic_reshaper.reshape(text)  # إعادة ترتيب الأحرف
    bidi_text = get_display(reshaped_text)  # تغيير اتجاه النص

    lines = bidi_text.split('\n')
    total_lines = len(lines)
    pdf_paths = []

    with pdf_lock:
        if total_lines <= lines_per_page:
            pdf_paths.append(pdf_path)
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font("Arial", "", font_path, uni=True)
            pdf.set_font("Arial", size=30)
            pdf.multi_cell(0, 10, bidi_text)
            pdf.output(pdf_path)
        else:
            part_number = 1
            for i in range(0, total_lines, lines_per_page):
                part_text = '\n'.join(lines[i:i + lines_per_page])
                part_pdf_path = pdf_path.replace('.pdf', f'_part{part_number}.pdf')
                pdf_paths.append(part_pdf_path)

                pdf = FPDF()
                pdf.add_page()
                pdf.add_font("Arial", "", font_path, uni=True)
                pdf.set_font("Arial", size=25)
                pdf.multi_cell(0, 10, part_text)
                pdf.output(part_pdf_path)

                part_number += 1

    return pdf_paths

printer_settings_cache = None

def get_printer_settings():
    global printer_settings_cache
    if printer_settings_cache is not None:
        return printer_settings_cache

    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute("SELECT ip, port FROM printer_settings LIMIT 1")
        result = cursor.fetchone()
        if result:
            printer_settings_cache = (result[0], result[1])
            return printer_settings_cache
        else:
            raise Exception("No printer settings found")
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        raise
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

printer_lock = Lock()

def send_pdf_to_printer(pdf_path, ip, port, timeout=5):
    with printer_lock:
        try:
            with open(pdf_path, 'rb') as f:
                pdf_data = f.read()
            print(ip, port)
            s = socket.socket()
            s.settimeout(timeout)  # تحديد مهلة الاتصال بـ 5 ثوانٍ
            s.connect((ip, port))
            
            # تقسيم البيانات إلى أجزاء صغيرة
            chunk_size = 4096
            for i in range(0, len(pdf_data), chunk_size):
                s.sendall(pdf_data[i:i + chunk_size])
            
            s.close()
            print(f"تم إرسال أمر الطباعة بنجاح: {pdf_path}")
            return True
        except socket.timeout:
            print("انتهت مهلة الاتصال.")
            return False
        except ConnectionRefusedError:
            print("الاتصال مرفوض: تأكد من أن الطابعة تعمل وأن المنفذ 9100 مفتوح.")
            return False
        except socket.error as e:
            print(f"خطأ في الاتصال: {e}")
            return False

print_queue = Queue()

def printer_worker():
    while True:
        pdf_path, ip, port = print_queue.get()
        send_pdf_to_printer(pdf_path, ip, port)
        print_queue.task_done()

# بدء عامل الطابور
Thread(target=printer_worker, daemon=True).start()


def parse_meal_details(meal_string):
    meals = meal_string.split('|')
    parsed_meals = []
    for meal in meals:
        qty, rest = meal.split(' x ')
        name, price = rest.split(' (')
        price = price.replace(' دينار)', '')
        parsed_meals.append(f"{name:<30} {qty:<5} {price:>10}")
    return "\n\n".join(parsed_meals)

@app.route('/make-order', methods=['GET', 'POST'])
@levels_required('admin', 'editor')
def makeOrder():
    form = OrderForm()
    if request.method == 'POST':
        print(request.form)  # طباعة القيم المرسلة للتأكد من أنها صحيحة

    if form.validate_on_submit():
        order_uuid = form.order_id.data
        order = form.order.data
        phone = form.phone.data
        name = form.name.data
        address = form.address.data
        address_description = form.address_description.data
        note = form.note.data
        zone = form.zone.data
        tocost = form.tocost.data
        discount = form.discount.data
        ordemname = session.get('username')
        statu = "new"
        now = datetime.now()
        date = now.strftime("%Y/%m/%d")
        time = now.strftime("%H:%M:%S")
        ordtrk = "قيد التجهيز"
        current_time2 = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if discount.endswith('%'):
            discount = discount.rstrip('%')
        try:
            discount = float(discount)  # تحويل القيمة إلى عدد إذا كانت صالحة
        except ValueError:
            return jsonify({"error": "Invalid discount value"}), 400

        db = None
        cursor = None

        try:
            db = get_db_connection()
            cursor = db.cursor()

            check_sql = """
            SELECT 
                (SELECT COUNT(*) FROM ceorder WHERE order_uuid = %s) as order_exists,
                (SELECT COUNT(*) FROM cecustomer WHERE Cphone = %s) as customer_exists
            """
            cursor.execute(check_sql, (order_uuid, phone))
            result = cursor.fetchone()
            order_exists = result[0]
            customer_exists = result[1]

            if order_exists > 0:
                return jsonify({"error": "الطلب مكرر ولا يمكن إضافته"}), 400

            if customer_exists == 0:
                add_customer_query = "INSERT INTO cecustomer (Cname, Cphone, Caddress, Cnote, Corder) VALUES (%s, %s, %s, %s, %s)"
                cursor.execute(add_customer_query, (name, phone, address, note, order))
                db.commit()

            cursor.execute("SELECT MAX(ordercode) FROM ceorder")
            result = cursor.fetchone()
            last_ordercode = result[0] if result[0] is not None else 0
            ordercode = last_ordercode + 1

            sql = """
            INSERT INTO ceorder 
            (ordercode, Cphone, Cname, `order`, location, locdesc, `date`, `time`, statu, note, oroutinzone, orddiscnt, ordemname, ordtocost, ordtracking, order_preparing_start_time, order_uuid) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            values = (ordercode, phone, name, order, address, address_description, date, time, statu, note, zone, discount, ordemname, tocost, ordtrk, current_time2, order_uuid)
            cursor.execute(sql, values)
            
            db.commit()
            
            order_details = get_order_details(ordercode)
            meal_details = parse_meal_details(order_details['order'])

            delivery_price = " " if order_details['oroutinzone'] == "داخل الزون" else "3000.00 دينار"

            order_receipt = f"""
        
******* CHICKEN EXPRESS *******
            فرع المنصور

Order #: ({ordercode})

Date: {order_details['date']}    Time: {order_details['time']}

=============================
Customer

Phone: {order_details['Cphone']}    Name: {order_details['Cname']}

Address: {order_details['location']}

=============================

Order Details:
----------------------------------------
Price                       Qty      Item
----------------------------------------

{meal_details}

----------------------------------------

Delivery: {order_details['oroutinzone']} ({delivery_price})

Total: {order_details['ordtocost']}

=============================

Notes: {order_details['note']}

=============================

Thank you for choosing CHICKEN EXPRESS!
***************************************
***************************************
***************************************
"""
            print(order_receipt)  # طباعة التفاصيل للتأكد من البيانات
            pdf_file_path = "restaurant_order.pdf"
            font_path = r"Arial.ttf"  # تأكد من أن ملف Arial.ttf موجود في المسار الصحيح
            pdf_paths = create_pdfs_from_text(order_receipt, pdf_file_path, font_path, lines_per_page=50)

            # جلب إعدادات الطابعة من قاعدة البيانات
            printer_ip, printer_port = get_printer_settings()
            
            print_status = True
            # إرسال كل جزء للطابعة بشكل متتابع دون تأخير
            for path in pdf_paths:
                success = send_pdf_to_printer(path, printer_ip, printer_port)
                if not success:
                    print_status = False

            # تسجيل العملية في جدول aclog
            action_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log_action(ordercode, action_time, session.get('username'), 'makeOrder')

            # حذف ملفات PDF المؤقتة بعد إرسالها للطابعة
            try:
                for path in pdf_paths:
                    os.remove(path)
            except Exception as e:
                print(f"Error while removing the temporary files: {e}")

            return jsonify({"message": "تم استلام الطلب بنجاح", "print_status": print_status}), 200
        except mysql.connector.Error as err:
            print(f"Database error: {err}")
            return jsonify({"error": "حدث خطأ أثناء معالجة الطلب"}), 500
        except Exception as e:
            print(f"Unexpected error: {e}")
            return jsonify({"error": "حدث خطأ غير متوقع أثناء معالجة الطلب"}), 500
        finally:
            if cursor:
                cursor.close()
            if db:
                db.close()

    return jsonify({"error": "طلب غير صالح"}), 400



@app.route('/submit_edit_order', methods=['POST'])
@levels_required('admin', 'editor')
def submit_edit_order():
    data = request.json
    order_code = data.get('order_code')
    order_details = data.get('order_details')
    phone = data.get('phone')
    name = data.get('name')
    address = data.get('address')
    notes = data.get('notes')
    zone = data.get('zone')
    total_price = data.get('total_price')
    discount = data.get('discount')
    username = session.get('username')
    print("Received data:", order_code, order_details, phone, name, address, notes, zone, discount, total_price, username)
    
    # إزالة علامة % من الخصم إذا كانت موجودة
    if discount.endswith('%'):
        discount = discount.rstrip('%')
    try:
        discount = float(discount)  # تحويل القيمة إلى عدد إذا كانت صالحة
    except ValueError:
        return jsonify({"error": "Invalid discount value"}), 400

    try:
        db = get_db_connection()
        cursor = db.cursor()

        sql = """UPDATE ceorder
                 SET Cphone=%s, Cname=%s, `order`=%s, location=%s, note=%s, oroutinzone=%s, orddiscnt=%s, ordemname=%s, ordtocost=%s
                 WHERE ordercode=%s"""
        values = (phone, name, order_details, address, notes, zone, discount, username, total_price, order_code)
        cursor.execute(sql, values)
        db.commit()
        
        # تسجيل العملية في جدول aclog
        action_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_action(order_code, action_time, username, 'submit_edit_order')
        
        cursor.close()
        db.close()

        order = get_order_details(order_code)
        meal_details = parse_meal_details(order['order'])

        delivery_price = " " if order['oroutinzone'] == "داخل الزون" else "3000.00 دينار"

        order_receipt = f"""
        
******* CHICKEN EXPRESS *******
            فرع المنصور
            * طلب معدل

Order #: ({order_code})

Date: {order['date']}    Time: {order['time']}

=============================
Customer

Phone: {order['Cphone']}    Name: {order['Cname']}

Address: {order['location']}

=============================

Order Details:
----------------------------------------
Price                       Qty      Item
----------------------------------------

{meal_details}

----------------------------------------

Delivery: {order['oroutinzone']} ({delivery_price})

Total: {order['ordtocost']}

=============================

Notes: {order['note']}

=============================

Thank you for choosing CHICKEN EXPRESS!
***************************************
***************************************
***************************************
"""
        pdf_file_path = "restaurant_order.pdf"
        font_path = r"Arial.ttf"
        pdf_paths = create_pdfs_from_text(order_receipt, pdf_file_path, font_path, lines_per_page=50)
        
        # جلب إعدادات الطابعة من قاعدة البيانات
        printer_ip, printer_port = get_printer_settings()

        # إرسال كل جزء للطابعة بشكل متتابع دون تأخير
        for path in pdf_paths:
            send_pdf_to_printer(path, printer_ip, printer_port)

        return jsonify({'success': True}), 200
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء معالجة الطلب'}), 500

@app.route('/print_order', methods=['POST'])
@levels_required('admin', 'editor')
def print_order():
    data = request.json
    ordercode = data.get('ordercode')
    
    if not ordercode:
        return jsonify({'success': False, 'error': 'رقم الطلب غير موجود'}), 400
    
    try:
        order = get_order_details(ordercode)
        meal_details = parse_meal_details(order['order'])

        delivery_price = " " if order['oroutinzone'] == "داخل الزون" else "3000.00 دينار"

        order_receipt = f"""
        
******* CHICKEN EXPRESS *******
            فرع المنصور
           معاد الطبع

Order #: ({ordercode})

Date: {order['date']}    Time: {order['time']}

=============================
Customer

Phone: {order['Cphone']}    Name: {order['Cname']}

Address: {order['location']}

=============================

Order Details:
----------------------------------------
Price                       Qty      Item
----------------------------------------

{meal_details}

----------------------------------------

Delivery: {order['oroutinzone']} ({delivery_price})

Total: {order['ordtocost']}

=============================

Notes: {order['note']}

=============================

Thank you for choosing CHICKEN EXPRESS!
***************************************
***************************************
***************************************
"""
        pdf_file_path = "restaurant_order.pdf"
        font_path = r"Arial.ttf"  # تأكد من أن ملف Arial.ttf موجود في المسار الصحيح
        pdf_paths = create_pdfs_from_text(order_receipt, pdf_file_path, font_path, lines_per_page=50)

        # جلب إعدادات الطابعة من قاعدة البيانات
        printer_ip, printer_port = get_printer_settings()
        
        # إرسال كل جزء للطابعة بشكل متتابع دون تأخير
        for path in pdf_paths:
            send_pdf_to_printer(path, printer_ip, printer_port)
        
        # تسجيل العملية في جدول aclog
        action_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_action(ordercode, action_time, session.get('username'), 'print_order')

        return jsonify({'success': True}), 200
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500





    

@cache.cached(timeout=3600, key_prefix='menu_cache')
def showMenu():
    try:
        db = get_db_connection()
        if db is None:
            return []
        query = "SELECT * FROM ceogba"
        cursor = db.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        db.close()
        return rows
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return []

@app.route('/static/<path:filename>')
@levels_required('admin', 'editor')
def static_files(filename):
    response = make_response(send_from_directory('static', filename))
    response.headers['Cache-Control'] = 'public, max-age=31536000'
    return response

export_status = {
    "status": "not_started",
    "file_path": ""
}

@app.route('/export', methods=['POST'])
@levels_required('admin', 'editor')
def start_export():
    try:
        data = request.json
        time_period = data.get('timePeriod')
        start_date = data.get('startDate')
        end_date = data.get('endDate')

        export_status["status"] = "in_progress"
        export_status["file_path"] = ""

        threading.Thread(target=export_data_task, args=(time_period, start_date, end_date)).start()
        return jsonify({'message': 'تم بدء عملية تصدير البيانات في الخلفية'}), 202
    except Exception as e:
        return jsonify({'error': f'حدث خطأ أثناء بدء عملية تصدير البيانات: {str(e)}'}), 500

def get_time_period_clause(time_period, start_date=None, end_date=None):
    if time_period == 'daily':
        return " AND DATE(`date`) = CURDATE()", []
    elif time_period == 'weekly':
        return " AND YEARWEEK(`date`, 1) = YEARWEEK(CURDATE(), 1)", []
    elif time_period == 'monthly':
        return " AND YEAR(`date`) = YEAR(CURDATE()) AND MONTH(`date`) = MONTH(CURDATE())", []
    elif time_period == 'custom':
        return " AND DATE(`date`) BETWEEN %s AND %s", [start_date, end_date]
    else:
        return "", []

def export_data_task(time_period, start_date, end_date):
    try:
        db = get_db_connection()
        cursor = db.cursor()

        query_base = """
            SELECT ordercode, Cphone, Cname, `order`, location, `date`, `time`, statu, note, oroutinzone, orddiscnt, ordemname, ordtocost, ordtracking, orddrivename,
                   order_preparing_start_time, order_preparing_end_time, order_delivery_start_time, order_delivered_time
            FROM ceorder
            WHERE 1=1
        """
        
        time_clause, time_params = get_time_period_clause(time_period, start_date, end_date)
        query_base += time_clause

        cursor.execute(query_base, time_params if time_params else ())
        data = cursor.fetchall()
        cursor.close()
        db.close()

        columns = ['Order Code', 'Phone', 'Name', 'Order', 'Location', 'Date', 'Time', 'Status', 'Note', 'Routing Zone', 'Discount', 'Employee Name', 'Total Cost', 'Order Tracking', 'Driver Name',
                   'Preparing Start Time', 'Preparing End Time', 'Delivery Start Time', 'Delivered Time', 'Preparing Duration', 'Driver Wait Duration', 'Delivery Duration', 'Total Order Duration']
        df = pd.DataFrame(data, columns=columns[:-4])

        def calculate_duration(start, end):
            if pd.isna(start) or pd.isna(end):
                return None
            duration = (pd.to_datetime(end) - pd.to_datetime(start)).total_seconds()
            return duration if duration >= 0 else None

        def format_duration(seconds):
            if seconds is None:
                return None
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            seconds = int(seconds % 60)
            return f"{hours:02}:{minutes:02}:{seconds:02}"

        df['Preparing Duration'] = df.apply(lambda row: format_duration(calculate_duration(row['Preparing Start Time'], row['Preparing End Time'])), axis=1)
        df['Driver Wait Duration'] = df.apply(lambda row: format_duration(calculate_duration(row['Preparing End Time'], row['Delivery Start Time'])), axis=1)
        df['Delivery Duration'] = df.apply(lambda row: format_duration(calculate_duration(row['Delivery Start Time'], row['Delivered Time'])), axis=1)
        df['Total Order Duration'] = df.apply(lambda row: format_duration(calculate_duration(row['Preparing Start Time'], row['Delivered Time'])), axis=1)

        if time_period == 'daily':
            file_name = 'daily_report.xlsx'
        elif time_period == 'weekly':
            file_name = 'weekly_report.xlsx'
        elif time_period == 'monthly':
            file_name = 'monthly_report.xlsx'
        elif time_period == 'custom':
            file_name = f'custom_report_{start_date}_to_{end_date}.xlsx'
        else:
            file_name = 'report.xlsx'
            
        file_path = os.path.join('exports', file_name)
        os.makedirs('exports', exist_ok=True)
        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active
        ws.title = "Orders"

        ws_stats = wb.create_sheet(title="Statistics 📊")
        ws_status = wb.create_sheet(title="Order Status 📝")
        ws_meals = wb.create_sheet(title="Meals Statistics 🍽️")
        ws_customers = wb.create_sheet(title="Top Customers 👥")
        ws_revenue = wb.create_sheet(title="Daily Revenue 💰")

        sheet_configs = {
            "Orders": {"color": "FFC000"},
            "Statistics 📊": {"color": "4DA3FF"},
            "Order Status 📝": {"color": "EE4848"},
            "Meals Statistics 🍽️": {"color": "955AC1"},
            "Top Customers 👥": {"color": "1BD059"},
            "Daily Revenue 💰": {"color": "FFFF00"}
        }

        for sheet_name, config in sheet_configs.items():
            sheet = wb[sheet_name]
            header_fill = PatternFill(start_color=config['color'], end_color=config['color'], fill_type="solid")
            header_font = Font(bold=True, color="000000")
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alternating_fill = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                            PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")]

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        driver_wait_duration_col = columns.index('Driver Wait Duration') + 1
        delivery_duration_col = columns.index('Delivery Duration') + 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            fill = alternating_fill[row_idx % 2]
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = fill
                if cell.column == 7:
                    cell.number_format = 'hh:mm:ss'
                if cell.column == 6:
                    cell.number_format = 'yyyy-mm-dd'

                if cell.column in [driver_wait_duration_col, delivery_duration_col] and cell.value:
                    try:
                        duration_seconds = sum(x * int(t) for x, t in zip([3600, 60, 1], map(int, cell.value.split(":"))))
                    except AttributeError:
                        duration_seconds = (cell.value.hour * 3600) + (cell.value.minute * 60) + cell.value.second
                    if duration_seconds >= 600:
                        cell.fill = red_fill

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    cell_value_length = len(str(cell.value))
                    if cell_value_length > max_length:
                        max_length = cell_value_length
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        def format_sheet(sheet, headers, data):
            header_fill = PatternFill(start_color=sheet_configs[sheet.title]['color'], end_color=sheet_configs[sheet.title]['color'], fill_type="solid")
            header_font = Font(bold=True, color="000000")
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font

            for row_idx, row_data in enumerate(data, start=2):
                fill = alternating_fill[row_idx % 2]
                for col_num, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_idx, column=col_num)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = fill

            for column_cells in sheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    if cell.value:
                        cell_value_length = len(str(cell.value))
                        if cell_value_length > max_length:
                            max_length = cell_value_length
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

        def get_query_results(query, params=None):
            db = get_db_connection()
            cursor = db.cursor()
            cursor.execute(query, params)
            data = cursor.fetchall()
            cursor.close()
            db.close()
            return data

        time_clause, time_params = get_time_period_clause(time_period, start_date, end_date)

        # الجزء الأول: حساب عدد الطلبات لكل ساعة لكل يوم
        query_orders_per_hour_per_day = """
            SELECT DATE(`date`) as day, HOUR(`time`) as hour, COUNT(*) as orders_count
            FROM ceorder
            WHERE 1=1
        """ + time_clause + """
            GROUP BY DATE(`date`), HOUR(`time`)
        """

        orders_per_hour_data_per_day = get_query_results(query_orders_per_hour_per_day, time_params)

        daily_data = {}
        for row in orders_per_hour_data_per_day:
            day, hour, count = row
            if day not in daily_data:
                daily_data[day] = [0] * 24
            daily_data[day][hour] = count

        stats_data = []
        for day, counts in daily_data.items():
            total_orders = sum(counts)
            day_str = day.strftime('%Y-%m-%d')
            hours_data = [day_str] + counts + [total_orders]
            stats_data.append(hours_data)

        headers = ['Day'] + [f"{hour % 12 or 12} {'AM' if hour < 12 else 'PM'}" for hour in range(24)] + ['Total Orders']

        # تنسيق العناوين وإضافة إطار للجداول
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alternating_fill = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                            PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")]

        # تعيين موقع الجدول الأول
        start_row_1 = 28
        for col_num, header in enumerate(headers, start=1):
            cell = ws_stats.cell(row=start_row_1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_idx, row_data in enumerate(stats_data, start=start_row_1 + 1):
            fill = alternating_fill[(row_idx - start_row_1 - 1) % 2]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_stats.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.fill = fill

        chart_orders_per_day = BarChart()
        data_per_day = Reference(ws_stats, min_col=2, min_row=start_row_1, max_row=start_row_1 + len(stats_data), max_col=25)
        cats_per_day = Reference(ws_stats, min_col=1, min_row=start_row_1 + 1, max_row=start_row_1 + len(stats_data))
        chart_orders_per_day.add_data(data_per_day, titles_from_data=True)
        chart_orders_per_day.set_categories(cats_per_day)
        chart_orders_per_day.title = "عدد الطلبات لكل ساعة لكل يوم"
        chart_orders_per_day.y_axis.title = 'عدد الطلبات'
        chart_orders_per_day.x_axis.title = 'اليوم'
        chart_orders_per_day.style = 11
        chart_orders_per_day.width = 20
        chart_orders_per_day.height = 10
        ws_stats.add_chart(chart_orders_per_day, "AB28")

        # الجزء الثاني: حساب عدد الطلبات لكل ساعة
        query_orders_per_hour = """
            SELECT HOUR(`time`) as hour, COUNT(*) as orders_count
            FROM ceorder
            WHERE 1=1
        """ + time_clause + """
            GROUP BY HOUR(`time`)
        """

        orders_per_hour_data = get_query_results(query_orders_per_hour, time_params)

        # التأكد من البيانات المسترجعة
        print("Orders per hour data:", orders_per_hour_data)

        # ساعات اليوم كاملة من 0 إلى 23
        all_hours = list(range(24))
        orders_counts_dict = {row[0]: row[1] for row in orders_per_hour_data}
        orders_counts = [orders_counts_dict.get(hour, 0) for hour in all_hours]
        hours_12 = [f"{hour % 12 or 12} {'صباحًا' if hour < 12 else 'مساءً'}" for hour in all_hours]

        # تعيين موقع الجدول الثاني
        start_row_2 = 1
        ws_stats.cell(row=start_row_2, column=1).value = '⏰ Hour'
        ws_stats.cell(row=start_row_2, column=2).value = 'Orders Count'

        # تنسيق العناوين وإضافة إطار للجداول
        for col_num, header in enumerate(['⏰ Hour', 'Orders Count'], start=1):
            cell = ws_stats.cell(row=start_row_2, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_idx, (hour, count) in enumerate(zip(hours_12, orders_counts), start=start_row_2 + 1):
            fill = alternating_fill[(row_idx - start_row_2 - 1) % 2]
            ws_stats.cell(row=row_idx, column=1).value = hour
            ws_stats.cell(row=row_idx, column=1).border = thin_border
            ws_stats.cell(row=row_idx, column=1).fill = fill

            ws_stats.cell(row=row_idx, column=2).value = count
            ws_stats.cell(row=row_idx, column=2).border = thin_border
            ws_stats.cell(row=row_idx, column=2).fill = fill

        chart_orders = BarChart()
        data = Reference(ws_stats, min_col=2, min_row=start_row_2, max_row=start_row_2 + len(orders_counts))
        cats = Reference(ws_stats, min_col=1, min_row=start_row_2 + 1, max_row=start_row_2 + len(hours_12))
        chart_orders.add_data(data, titles_from_data=True)
        chart_orders.set_categories(cats)
        chart_orders.title = "عدد الطلبات لكل ساعة"
        chart_orders.y_axis.title = 'عدد الطلبات'
        chart_orders.x_axis.title = 'الساعة'
        chart_orders.style = 11
        chart_orders.width = 20
        chart_orders.height = 10
        ws_stats.add_chart(chart_orders, "D1")


        # حالة الطلبات
        query_orders_status = """
            SELECT `statu`, COUNT(*)
            FROM ceorder
            WHERE 1=1
        """ + time_clause + """
            GROUP BY `statu`
        """
        orders_status_data = get_query_results(query_orders_status, time_params)

        status_translation = {
            'Done': 'طلبات تم توصيلها',
            'new': 'الطلبات الجديدة',
            'Lost': 'طلبات تم الغائها',
            
        }

        # الوان والعياذ بالله
        status_colors = {
            'طلبات تم توصيلها': '91cc75',  
            'الطلبات الجديدة': '5470c6',   
            'طلبات تم الغائها': 'ff0000', 

        }

        translated_orders_status_data = [(status_translation.get(row[0], row[0]), row[1]) for row in orders_status_data]

        # استخراج الملصقات والمجموعات بعد الترجمة
        status_labels = [row[0] for row in translated_orders_status_data]
        status_counts = [row[1] for row in translated_orders_status_data]
        total_orders = sum(status_counts)

        format_sheet(ws_status, ['📊 Status', 'Count'], zip(status_labels, status_counts))
        ws_status.append(["إجمالي الطلبات", total_orders])

        chart_status = PieChart()
        data = Reference(ws_status, min_col=2, min_row=1, max_row=len(status_counts) + 1)
        labels = Reference(ws_status, min_col=1, min_row=2, max_row=len(status_labels) + 1)
        chart_status.add_data(data, titles_from_data=True)
        chart_status.set_categories(labels)
        chart_status.title = "حالة الطلبات"
        chart_status.height = 10
        chart_status.width = 20
        chart_status.legend.position = 'r'
        chart_status.dataLabels = DataLabelList()
        chart_status.dataLabels.showVal = True
        chart_status.dataLabels.showPercent = True


        for i, label in enumerate(status_labels):
            color = status_colors.get(label, 'FFFFFF')  # اللون الافتراضي هو أبيض إذا لم يتم تحديد اللون
            slice = DataPoint(idx=i)
            slice.graphicalProperties = GraphicalProperties(solidFill=ColorChoice(srgbClr=color))
            chart_status.series[0].data_points.append(slice)

        ws_status.add_chart(chart_status, "D1")


        # إحصائيات الوجبات
        query_top_meals = """
            SELECT `order`
            FROM ceorder
            WHERE 1=1
        """ + time_clause
        meals_data = get_query_results(query_top_meals, time_params)

        meal_counter = Counter()
        meal_quantity_counter = Counter()

        for row in meals_data:
            order_details = row[0]
            meals = order_details.split('|')
            for meal in meals:
                try:
                    quantity, meal_name = meal.split(' x ')
                    meal_name = meal_name.split('(')[0].strip()
                    quantity = int(quantity.strip())
                    if meal_name:
                        meal_counter[meal_name] += 1
                        meal_quantity_counter[meal_name] += quantity
                except ValueError:
                    continue

        # الجزء الأول: أكثر 6 وجبات متكررة حسب اسم الوجبة
        top_meals = meal_counter.most_common(6)
        meal_labels = [meal[0] for meal in top_meals]
        meal_counts = [meal[1] for meal in top_meals]

        format_sheet(ws_meals, ['🍔 Meal', 'Count'], zip(meal_labels, meal_counts))

        chart_meals = PieChart()
        data = Reference(ws_meals, min_col=2, min_row=1, max_row=len(meal_counts)+1)
        labels = Reference(ws_meals, min_col=1, min_row=2, max_row=len(meal_labels)+1)
        chart_meals.add_data(data, titles_from_data=True)
        chart_meals.set_categories(labels)
        chart_meals.title = "أكثر 6 وجبات متكررة"
        chart_meals.height = 10
        chart_meals.width = 20
        chart_meals.legend.position = 'r'
        chart_meals.dataLabels = DataLabelList()
        chart_meals.dataLabels.showVal = True
        chart_meals.dataLabels.showPercent = True
        ws_meals.add_chart(chart_meals, "D1")

        # الجزء الثاني: جميع الوجبات وكمية الطلبات لكل وجبة
        all_meals_data = [(meal, count) for meal, count in meal_quantity_counter.items()]

        # فرز البيانات بناءً على الكميات المطلوبة من الأكثر إلى الأقل
        all_meals_data = sorted(all_meals_data, key=lambda x: x[1], reverse=True)

        # حساب المركز المناسب للجدول الجديد
        new_start_row = ws_meals.max_row + 5  # بدء الجدول الجديد بعد 5 صفوف من آخر صف موجود
        new_start_col = 1  # بدء الجدول من العمود الأول

        # إدراج البيانات في الجدول الجديد
        ws_meals.cell(row=new_start_row, column=new_start_col, value="اسم الوجبة")
        ws_meals.cell(row=new_start_row, column=new_start_col + 1, value="عدد الطلبات")

        header_fill = PatternFill(start_color="955AC1", end_color="955AC1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        ws_meals.cell(row=new_start_row, column=new_start_col).fill = header_fill
        ws_meals.cell(row=new_start_row, column=new_start_col).font = header_font
        ws_meals.cell(row=new_start_row, column=new_start_col + 1).fill = header_fill
        ws_meals.cell(row=new_start_row, column=new_start_col + 1).font = header_font

        alternating_fill = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                            PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")]

        for idx, (meal_name, count) in enumerate(all_meals_data, start=new_start_row + 1):
            ws_meals.cell(row=idx, column=new_start_col, value=meal_name)
            ws_meals.cell(row=idx, column=new_start_col + 1, value=count)

            fill = alternating_fill[(idx - new_start_row) % 2]  # Ensure proper alternation
            ws_meals.cell(row=idx, column=new_start_col).fill = fill
            ws_meals.cell(row=idx, column=new_start_col + 1).fill = fill

        # تنسيق الحدود والتوسيط
        for row in ws_meals.iter_rows(min_row=new_start_row, max_row=new_start_row + len(all_meals_data), min_col=new_start_col, max_col=new_start_col + 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # تعديل عرض الأعمدة
        for col in range(new_start_col, new_start_col + 2):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws_meals[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_meals.column_dimensions[column_letter].width = adjusted_width



        # أفضل الزبائن
        query_top_customers = """
            SELECT Cname, Cphone, COUNT(*) as orders
            FROM ceorder
            WHERE `statu` = 'Done'
        """ + time_clause + """
            GROUP BY Cname, Cphone
            ORDER BY orders DESC
            LIMIT 10
        """
        top_customers_data = get_query_results(query_top_customers, time_params)

        customer_labels = [f"{row[0]} ({row[1]})" for row in top_customers_data]
        customer_orders = [row[2] for row in top_customers_data]

        format_sheet(ws_customers, ['🏆 Customer', 'Orders'], zip(customer_labels, customer_orders))

        chart_customers = BarChart()
        data = Reference(ws_customers, min_col=2, min_row=1, max_row=len(customer_orders)+1)
        labels = Reference(ws_customers, min_col=1, min_row=2, max_row=len(customer_labels)+1)
        chart_customers.add_data(data, titles_from_data=True)
        chart_customers.set_categories(labels)
        chart_customers.title = "أكثر الزبائن طلبًا"
        chart_customers.height = 10
        chart_customers.width = 20
        chart_customers.legend.position = 'r'
        ws_customers.add_chart(chart_customers, "D1")

        # أرباح
        query_daily_revenue = """
            SELECT SUM(ordtocost)
            FROM ceorder
            WHERE `statu` = 'Done'
        """ + time_clause
        daily_revenue_data = get_query_results(query_daily_revenue, time_params)
        daily_revenue = daily_revenue_data[0][0] if daily_revenue_data[0][0] else 0

        if time_period == 'daily':
            date = datetime.now().strftime('%Y-%m-%d')
        elif time_period == 'weekly':
            start_of_week = (datetime.now() - timedelta(days=datetime.now().weekday())).strftime('%Y-%m-%d')
            end_of_week = (datetime.now() + timedelta(days=(6 - datetime.now().weekday()))).strftime('%Y-%m-%d')
            date = f"Week: {start_of_week} to {end_of_week}"
        elif time_period == 'monthly':
            date = datetime.now().strftime('%Y-%m')
        elif time_period == 'custom':
            date = f"{start_date} to {end_date}"

        format_sheet(ws_revenue, ['Date', '💰 Revenue'], [(date, daily_revenue)])

        chart_revenue = BarChart()
        data = Reference(ws_revenue, min_col=2, min_row=1, max_row=2)
        cats = Reference(ws_revenue, min_col=1, min_row=2, max_row=2)
        chart_revenue.add_data(data, titles_from_data=True)
        chart_revenue.set_categories(cats)
        chart_revenue.title = "الأرباح"
        chart_revenue.y_axis.title = 'Revenue'
        chart_revenue.x_axis.title = 'Date'
        chart_revenue.style = 11
        chart_revenue.width = 20
        chart_revenue.height = 10
        ws_revenue.add_chart(chart_revenue, "D1")


        wb.save(file_path)

        export_status["status"] = "completed"
        export_status["file_path"] = file_path
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        export_status["status"] = "failed"

@app.route('/export/status', methods=['GET'])
def export_status_check():
    return jsonify(export_status), 200

@app.route('/export/download', methods=['GET'])
def download_file():
    if export_status["status"] == "completed" and os.path.exists(export_status["file_path"]):
        return send_file(export_status["file_path"], as_attachment=True)
    else:
        return jsonify({'error': 'الملف غير جاهز أو حدث خطأ أثناء التصدير'}), 404
    
    
# شييييت سواق 
export_driver_status = {
    "status": "not_started",
    "file_path": ""
}

@app.route('/export_driver', methods=['POST'])
def start_export_driver():
    try:
        data = request.json
        time_period = data.get('timePeriod')
        start_date = data.get('startDate')
        end_date = data.get('endDate')

        export_driver_status["status"] = "in_progress"
        export_driver_status["file_path"] = ""

        threading.Thread(target=export_driver_data_task, args=(time_period, start_date, end_date)).start()
        return jsonify({'message': 'تم بدء عملية تصدير البيانات الخاصة بالسائقين في الخلفية'}), 202
    except Exception as e:
        return jsonify({'error': f'حدث خطأ أثناء بدء عملية تصدير البيانات الخاصة بالسائقين: {str(e)}'}), 500

def export_driver_data_task(time_period, start_date, end_date):
    try:
        db = get_db_connection()
        cursor = db.cursor()

        query_base = """
            SELECT ordercode, Cphone, location, `date`, `time`, statu, orddrivename,
                   order_preparing_end_time, order_delivery_start_time, order_delivered_time
            FROM ceorder
            WHERE 1=1
        """
        
        time_clause, time_params = get_time_period_clause(time_period, start_date, end_date)
        query_base += time_clause

        cursor.execute(query_base, time_params if time_params else ())
        data = cursor.fetchall()
        cursor.close()
        db.close()

        columns = ['Order Code', 'Phone', 'Location', 'Date', 'Time', 'Status', 'Driver Name', 'Preparing End Time', 'Delivery Start Time', 'Delivered Time', 'Driver Wait Duration', 'Delivery Duration']
        df = pd.DataFrame(data, columns=columns[:-2])

        # Convert 'Time' column to timedelta and then to string in HH:MM:SS format
        df['Time'] = pd.to_timedelta(df['Time']).apply(lambda x: f"{x.components.hours:02}:{x.components.minutes:02}:{x.components.seconds:02}")

        def calculate_duration(start, end):
            if pd.isna(start) or pd.isna(end):
                return None
            duration = (pd.to_datetime(end) - pd.to_datetime(start)).total_seconds()
            return duration if duration >= 0 else None

        def format_duration(seconds):
            if seconds is None:
                return None
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            seconds = int(seconds % 60)
            return f"{hours:02}:{minutes:02}:{seconds:02}"

        df['Driver Wait Duration'] = df.apply(lambda row: format_duration(calculate_duration(row['Preparing End Time'], row['Delivery Start Time'])), axis=1)
        df['Delivery Duration'] = df.apply(lambda row: format_duration(calculate_duration(row['Delivery Start Time'], row['Delivered Time'])), axis=1)

        file_name = 'driver_report.xlsx'
        file_path = os.path.join('exports', file_name)
        os.makedirs('exports', exist_ok=True)
        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active
        ws.title = "Driver Orders"

        # Custom styling
        ws_stats = wb.create_sheet(title="Orders Stats 📊")
        ws_driver = wb.create_sheet(title="Driver Delays 🚗")

        sheet_configs = {
            "Driver Orders": {"color": "FFC000"},
            "Orders Stats 📊": {"color": "4DA3FF"},
            "Driver Delays 🚗": {"color": "EE4848"}
        }

        for sheet_name, config in sheet_configs.items():
            sheet = wb[sheet_name]
            header_fill = PatternFill(start_color=config['color'], end_color=config['color'], fill_type="solid")
            header_font = Font(bold=True, color="000000")
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alternating_fill = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                            PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")]

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        driver_wait_duration_col = columns.index('Driver Wait Duration') + 1
        delivery_duration_col = columns.index('Delivery Duration') + 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            fill = alternating_fill[row_idx % 2]
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = fill
                if cell.column == 5:  # Column 'Time'
                    cell.number_format = 'HH:MM:SS'
                if cell.column == 4:  # Column 'Date'
                    cell.number_format = 'yyyy-mm-dd'

                # Check and apply red fill for 'Driver Wait Duration' and 'Delivery Duration'
                if cell.column in [driver_wait_duration_col, delivery_duration_col] and cell.value:
                    try:
                        duration_seconds = sum(x * int(t) for x, t in zip([3600, 60, 1], map(int, cell.value.split(":"))))
                    except AttributeError:
                        duration_seconds = (cell.value.hour * 3600) + (cell.value.minute * 60) + cell.value.second
                    if duration_seconds >= 600:  # 10 minutes in seconds
                        cell.fill = red_fill

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    cell_value_length = len(str(cell.value))
                    if cell_value_length > max_length:
                        max_length = cell_value_length
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        def format_sheet(sheet, headers, data):
            header_fill = PatternFill(start_color=sheet_configs[sheet.title]['color'], end_color=sheet_configs[sheet.title]['color'], fill_type="solid")
            header_font = Font(bold=True, color="000000")
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font

            for row_idx, row_data in enumerate(data, start=2):
                fill = alternating_fill[row_idx % 2]
                for col_num, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_idx, column=col_num)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = fill

            for column_cells in sheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    if cell.value:
                        cell_value_length = len(str(cell.value))
                        if cell_value_length > max_length:
                            max_length = cell_value_length
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

        def get_query_results(query, params=None):
            db = get_db_connection()
            cursor = db.cursor()
            cursor.execute(query, params)
            data = cursor.fetchall()
            cursor.close()
            db.close()
            return data

        # بيانات ساعات الطلبات
        time_clause, time_params = get_time_period_clause(time_period, start_date, end_date)

        # عدد الطلبات لكل ساعة
        query_orders_per_hour = """
            SELECT HOUR(`time`) as hour, COUNT(*) as orders_count
            FROM ceorder
            WHERE 1=1
        """ + time_clause + """
            GROUP BY HOUR(`time`)
        """
        orders_per_hour_data = get_query_results(query_orders_per_hour, time_params)

        # ساعات اليوم كاملة من 0 إلى 23
        all_hours = list(range(24))
        orders_counts_dict = {row[0]: row[1] for row in orders_per_hour_data}
        orders_counts = [orders_counts_dict.get(hour, 0) for hour in all_hours]
        hours_12 = [f"{hour % 12 or 12} {'AM' if hour < 12 else 'PM'}" for hour in all_hours]

        format_sheet(ws_stats, ['⏰ Hour', 'Orders Count'], zip(hours_12, orders_counts))

        # إنشاء رسم بياني لعدد الطلبات لكل ساعة
        chart_orders = BarChart()
        data = Reference(ws_stats, min_col=2, min_row=1, max_row=len(orders_counts)+1)
        cats = Reference(ws_stats, min_col=1, min_row=2, max_row=len(hours_12)+1)
        chart_orders.add_data(data, titles_from_data=True)
        chart_orders.set_categories(cats)
        chart_orders.title = "عدد الطلبات لكل ساعة"
        chart_orders.y_axis.title = 'عدد الطلبات'
        chart_orders.x_axis.title = 'الساعة'
        chart_orders.style = 11
        chart_orders.width = 20
        chart_orders.height = 10
        ws_stats.add_chart(chart_orders, "D1")

        # تأخيرات السائقين
        def average_duration(duration_series):
            total_seconds = duration_series.dropna().apply(lambda x: sum(int(t) * 60 ** i for i, t in enumerate(reversed(x.split(":"))))).sum()
            return format_duration(total_seconds / len(duration_series)) if len(duration_series) > 0 else None

        driver_delays = df.groupby('Driver Name').agg(
            Avg_Wait_Duration=('Driver Wait Duration', average_duration),
            Avg_Delivery_Duration=('Delivery Duration', average_duration)
        ).reset_index()

        format_sheet(ws_driver, ['Driver Name', 'Avg Wait Duration', 'Avg Delivery Duration'], driver_delays.values)

        chart_driver_delays = BarChart()
        data = Reference(ws_driver, min_col=2, min_row=1, max_col=3, max_row=len(driver_delays) + 1)
        cats = Reference(ws_driver, min_col=1, min_row=2, max_row=len(driver_delays) + 1)
        chart_driver_delays.add_data(data, titles_from_data=True)
        chart_driver_delays.set_categories(cats)
        chart_driver_delays.title = "Driver Delays"
        chart_driver_delays.y_axis.title = 'Average Duration'
        chart_driver_delays.x_axis.title = 'Driver'
        chart_driver_delays.style = 11
        chart_driver_delays.width = 20
        chart_driver_delays.height = 10
        ws_driver.add_chart(chart_driver_delays, "E1")

        wb.save(file_path)

        export_driver_status["status"] = "completed"
        export_driver_status["file_path"] = file_path
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        export_driver_status["status"] = "failed"

        wb.save(file_path)

        export_driver_status["status"] = "completed"
        export_driver_status["file_path"] = file_path
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        export_driver_status["status"] = "failed"

@app.route('/export_driver/status', methods=['GET'])
def export_driver_status_check():
    return jsonify(export_driver_status), 200

@app.route('/export_driver/download', methods=['GET'])
def download_driver_file():
    if export_driver_status["status"] == "completed" and os.path.exists(export_driver_status["file_path"]):
        return send_file(export_driver_status["file_path"], as_attachment=True)
    else:
        return jsonify({'error': 'الملف غير جاهز أو حدث خطأ أثناء التصدير'}), 404



def convert_timedelta(obj):
    if isinstance(obj, timedelta):
        return str(obj)
    raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")

def calculate_total_cost(order_details):
    total_cost = 0
    items = order_details.split('|')
    for item in items:
        match = re.match(r'(\d+)\s*x\s*.*?\((\d+(\.\d+)?)\s*دينار\)', item)
        if match:
            price = float(match.group(2))
            total_cost += price
    return total_cost

def convert_timedelta(td):
    return str(td)

@app.route('/checkPhone', methods=['POST'])
@levels_required('admin', 'editor')
def check_phone():
    data = request.get_json()
    phone = data.get('phone')
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        query = """
        SELECT Cphone, Cname, `order`, location, `date`, `time`, statu, note, oroutinzone, orddiscnt
        FROM ceorder
        WHERE Cphone = %s AND `order` = (
            SELECT `order`
            FROM ceorder
            WHERE Cphone = %s
            GROUP BY `order`
            ORDER BY COUNT(*) DESC
            LIMIT 1
        )
        """
        cursor.execute(query, (phone, phone))
        results = cursor.fetchall()
        cursor.close()
        db.close()

        for result in results:
            result['ordtocost'] = calculate_total_cost(result['order'])
            for key, value in result.items():
                if isinstance(value, timedelta):
                    result[key] = convert_timedelta(value)

        return jsonify(results)
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return jsonify({'error': 'حدث خطأ أثناء التحقق من رقم الهاتف'}), 500

@app.route('/get_order_by_code', methods=['POST'])
def get_order_by_code_endpoint():
    order_code = request.json.get('order_code')
    data = get_order_by_code(order_code)
    return jsonify(data)


def get_order_by_code(order_code):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        query = "SELECT ordercode, Cphone, Cname, `order`, location, `date`, `time`, statu, note, oroutinzone, orddiscnt, ordemname, ordtocost FROM ceorder WHERE ordercode = %s"
        cursor.execute(query, (order_code,))
        row = cursor.fetchone()

        data = {}
        if row:
            order_date = row[5]
            order_time = row[6]

            if isinstance(order_time, timedelta):
                total_seconds = order_time.total_seconds()
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                order_time = time(hours, minutes)

            order_date_time = datetime.combine(order_date, order_time)
            current_time = datetime.now()
            time_difference = current_time - order_date_time

            if time_difference > timedelta(minutes=50):
                data['error'] = "لا يمكن تعديل الطلب، تجاوزت مدة الطلب 3 دقائق"
            elif row[7] != 'new':
                data['error'] = "لا يمكن تعديل الطلب، حالة الطلب ليست none"
            else:
                data = {
                    'ordercode': row[0],
                    'phone': row[1],
                    'name': row[2],
                    'order': row[3],
                    'address': row[4],
                    'status': row[7],
                    'note': row[8],
                    'zone': row[9],
                    'discount': row[10],
                    'employee_name': row[11],
                    'total_cost': row[12]
                }

        cursor.close()
        db.close()

        return data
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return {}
    
@app.route('/update_order_status', methods=['POST'])
@levels_required('admin', 'editor')
def update_order_status():
    data = request.json
    ordercode = data.get('ordercode')
    status = data.get('status')
    comment = data.get('comment', 'لا يوجد')
    wlost_reason = data.get('wlost', '')

    if not ordercode or not status:
        return jsonify({'success': False, 'error': 'البيانات غير كاملة'}), 400

    try:
        db = get_db_connection()
        cursor = db.cursor()

        cursor.execute("SELECT ordtracking, orddrivename FROM ceorder WHERE ordercode = %s", (ordercode,))
        order = cursor.fetchone()
        if status == 'Lost' and wlost_reason == 'إلغاء قبل خروج السائق' and order[0] == 'جار التسليم':
            return jsonify({'success': False, 'error': 'تم إرسال السائق. السائق: ' + order[1]}), 400

        if status == 'Lost':
            update_query = """
            UPDATE ceorder 
            SET statu = %s, wlost = %s, notelost = %s 
            WHERE ordercode = %s
            """
            cursor.execute(update_query, (status, wlost_reason, comment, ordercode))
        else:
            update_query = """
            UPDATE ceorder 
            SET statu = %s, note = %s 
            WHERE ordercode = %s
            """
            cursor.execute(update_query, (status, comment, ordercode))

        if status == 'تم التوصيل':
            update_query = """
            UPDATE ceorder 
            SET order_delivered_time = NOW() 
            WHERE ordercode = %s
            """
            cursor.execute(update_query, (ordercode,))

        db.commit()

        # تسجيل العملية في جدول aclog
        action_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_action(ordercode, action_time, session.get('username'), 'update_order_status')

        cursor.close()
        db.close()

        return jsonify({'success': True}), 200
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        return jsonify({'success': False, 'error': 'حدث خطأ أثناء تحديث حالة الطلب'}), 500
    
@app.route('/update_order_tracking', methods=['POST'])
@levels_required('admin', 'editor')
def update_order_tracking():
    data = request.json
    ordercode = data.get('ordercode')
    status = data.get('status')
    detail = data.get('detail', '')

    if not ordercode or not status:
        return jsonify({'success': False, 'error': 'البيانات غير كاملة'}), 400

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        db = get_db_connection()
        cursor = db.cursor()

        # استرجاع معلومات الطلب الحالية من قاعدة البيانات
        cursor.execute("""
            SELECT ordtracking, order_preparing_end_time, order_delivery_start_time, order_delivered_time 
            FROM ceorder 
            WHERE ordercode = %s
        """, (ordercode,))
        order_info = cursor.fetchone()

        if order_info:
            current_status = order_info[0]
            preparing_end_time = order_info[1]
            delivery_start_time = order_info[2]
            delivered_time = order_info[3]

            if status == "Prepared" and current_status != "تم التجهيز":
                if not preparing_end_time:
                    update_query = """
                    UPDATE ceorder 
                    SET order_preparing_end_time = %s, ordtracking = %s
                    WHERE ordercode = %s
                    """
                    cursor.execute(update_query, (timestamp, 'تم التجهيز', ordercode))
                else:
                    update_query = """
                    UPDATE ceorder 
                    SET ordtracking = %s
                    WHERE ordercode = %s
                    """
                    cursor.execute(update_query, ('تم التجهيز', ordercode))
            
            elif status == "InDelivery" and current_status != "جار التسليم":
                if not delivery_start_time:
                    update_query = """
                    UPDATE ceorder 
                    SET order_delivery_start_time = %s, orddrivename = %s, ordtracking = %s
                    WHERE ordercode = %s
                    """
                    cursor.execute(update_query, (timestamp, detail, 'جار التسليم', ordercode))
                else:
                    update_query = """
                    UPDATE ceorder 
                    SET orddrivename = %s, ordtracking = %s
                    WHERE ordercode = %s
                    """
                    cursor.execute(update_query, (detail, 'جار التسليم', ordercode))
            
            elif status == "Delivered" and current_status != "تم التوصيل":
                if not delivered_time:
                    update_query = """
                    UPDATE ceorder 
                    SET order_delivered_time = %s, statu = 'Done', ordtracking = %s
                    WHERE ordercode = %s
                    """
                    cursor.execute(update_query, (timestamp, 'تم التوصيل', ordercode))
                else:
                    update_query = """
                    UPDATE ceorder 
                    SET statu = 'Done', ordtracking = %s
                    WHERE ordercode = %s
                    """
                    cursor.execute(update_query, ('تم التوصيل', ordercode))
            
            db.commit()

            # تسجيل العملية في جدول aclog
            log_action(ordercode, timestamp, session.get('username'), 'update_order_tracking')
        
        cursor.close()
        db.close()

        return jsonify({'success': True}), 200
    except mysql.connector.Error as err:
        print("Database error:", err)
        return jsonify({'success': False, 'error': str(err)}), 500
    
@app.route('/update_driver_and_time', methods=['POST'])
@levels_required('admin', 'editor')
def update_driver_and_time():
    data = request.json
    ordercode = data.get('ordercode')
    new_driver_name = data.get('new_driver_name')

    if not ordercode or not new_driver_name:
        return jsonify({'success': False, 'error': 'البيانات غير كاملة'}), 400

    try:
        db = get_db_connection()
        cursor = db.cursor()

        cursor.execute("SELECT orddrivename FROM ceorder WHERE ordercode = %s", (ordercode,))
        order = cursor.fetchone()
        if not order:
            return jsonify({'success': False, 'error': 'الطلب غير موجود'}), 404

        old_driver_name = order[0]
        updated_driver_name = f"تم استبدال السائق: {old_driver_name} بالسائق: {new_driver_name}"
        update_query = """
        UPDATE ceorder
        SET orddrivename = %s, lostime = NOW(), wlost = %s
        WHERE ordercode = %s
        """
        cursor.execute(update_query, (updated_driver_name, 'حادث وجار تحويل الطلب', ordercode))

        db.commit()
        cursor.close()
        db.close()

        return jsonify({'success': True}), 200
    except mysql.connector.Error as err:
        print("Database error:", err)
        return jsonify({'success': False, 'error': str(err)}), 500




@app.route('/check_order_tracking', methods=['POST'])
def check_order_tracking():
    data = request.json
    ordercode = data.get('ordercode')

    if not ordercode:
        return jsonify({'success': False, 'error': 'البيانات غير كاملة'}), 400

    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute("SELECT ordtracking, orddrivename FROM ceorder WHERE ordercode = %s", (ordercode,))
        order = cursor.fetchone()
        db.close()

        if order[0] == 'جار التسليم':
            return jsonify({'can_cancel': False, 'driver_name': order[1]}), 200
        else:
            return jsonify({'can_cancel': True}), 200
    except mysql.connector.Error as err:
        print("Database error:", err)
        return jsonify({'success': False, 'error': str(err)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=8082, host='0.0.0.0')